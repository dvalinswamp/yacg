import os
from openpyxl import load_workbook
from PE import PE
from UPE import UPE


def SetLocation():
    outdir = "outputs/"
    if (not os.path.exists(outdir)):
        os.makedirs(outdir)


def ParseDevices():
    sheet = data.get_sheet_by_name('Devices')
    for row in range(2, sheet.max_row + 1):
        host = sheet['A' + str(row).strip()].value
        if (not host): print('Device undefinedint row', row)
        devtype = sheet['B' + str(row).strip()].value
        if (not devtype): print('Device ' + host + ' type absent')
        mgmtiface = sheet['C' + str(row).strip()].value
        if (not mgmtiface): print('Device ' + host + ' mgmtiface absent')
        ip = sheet['D' + str(row).strip()].value
        if (not ip): print('Device ' + host + ' mgmt ip absent')
        os = sheet['E' + str(row).strip()].value
        if (not os): print('Device ' + host + ' os definition absent')
        if (devtype == 'PE'):
            devices[host] = (PE(host, mgmtiface, ip, os))
        elif (devtype == 'UPE'):
            devices[host] = (UPE(host, mgmtiface, ip, os))
        else:
            print(devtype, ": devtype unknown")


def ParseInterfaces():
    sheet = data.get_sheet_by_name('Interfaces')
    for row in range(2, sheet.max_row + 1):
        device = str(sheet['A' + str(row).strip()].value)
        name = str(sheet['B' + str(row).strip()].value)
        media = str(sheet['C' + str(row).strip()].value)
        speed = str(sheet['D' + str(row).strip()].value)
        tag = str(sheet['E' + str(row).strip()].value)
        vrf = str(sheet['F' + str(row).strip()].value)
        ipv4_address = str(sheet['G' + str(row).strip()].value)
        ipv6_address = str(sheet['H' + str(row).strip()].value)
        description =  str(sheet['I' + str(row).strip()].value)
        devices[device].CreateInterface(name, speed, media, tag, vrf, ipv4_address, ipv6_address)
        print('created interface',name,'for pe',device)


def ParseVlans():
    sheet = data.get_sheet_by_name('Vlans')
    for row in range(2, sheet.max_row + 1):
        print('starting Vlans')
        device = sheet['A' + str(row).strip()].value
        vid = sheet['B' + str(row).strip()].value
        name = sheet['C' + str(row).strip()].value
        descr = sheet['D' + str(row).strip()].value
        print('processing', device, vid, name, descr)
        devices[device].CreateVlan(vid, name, descr)


def ParseL2Ifaces():
    sheet = data.get_sheet_by_name('L2ifaces')
    for row in range(2, sheet.max_row + 1):
        print('starting l2 interfaces')
        device = sheet['A' + str(row).strip()].value
        ifname = sheet['B' + str(row).strip()].value
        description =  sheet['C' + str(row).strip()].value
        mode = sheet['D' + str(row).strip()].value
        allowed_vlans = str(sheet['E' + str(row).strip()].value)
        print ('processing', device, ifname, mode, allowed_vlans)
        if (mode=='trunk'):
            print('creating a trunk from main')
            devices[device].CreateTrunk(ifname, description, allowed_vlans)
        elif (mode=='access'):
            devices[device].CreateAcc(ifname, description, allowed_vlans)

def ParseAE():
    sheet = data.get_sheet_by_name('AE')
    for row in range(2, sheet.max_row + 1):
        print('starting AE interfaces')
        device = sheet['A' + str(row).strip()].value
        abbr = sheet['B' + str(row).strip()].value
        aenum = sheet['C' + str(row).strip()].value
        devices[device].CreateAE(abbr, aenum)

def ParseVRF():
    sheet = data.get_sheet_by_name('VRF')
    for row in range(2, sheet.max_row + 1):
        print('starting VRFs')
        device = sheet['A' + str(row).strip()].value
        vrfname = sheet['B' + str(row).strip()].value
        rd = sheet['C' + str(row).strip()].value
        ipv4_import_RT = sheet['D' + str(row).strip()].value
        ipv4_export_RT = sheet['E' + str(row).strip()].value
        ipv6_import_RT = sheet['F' + str(row).strip()].value
        ipv6_export_RT = sheet['G' + str(row).strip()].value
        devices[device].CreateVRF(vrfname, rd, ipv4_import_RT, ipv4_export_RT, ipv6_import_RT, ipv6_export_RT)

def ParseLIB():
    sheet = data.get_sheet_by_name('LIB')
    for row in range(2, sheet.max_row+1):
        device = sheet['A' + str(row)].value
        iface = sheet['B' + str(row)].value
        labelproto = sheet['C' + str(row)].value
        igp = sheet['E' + str(row)].value
        area = sheet['F' + str(row)].value
        devices[device].CreateLabelledInterface(iface, labelproto, igp)
        devices[device].CreateIGPInterface(iface, igp, area)


SetLocation()

devices = {}
data = load_workbook('data.xlsx')
shits = data.get_sheet_names()

if ('Devices' in shits):
    print('Devices shit found')
    ParseDevices()
if ('Interfaces' in shits):
    print('Interfaces shit found')
    ParseInterfaces()
if ('Vlans' in shits):
    print('Vlans shit found')
    ParseVlans()
if ('AE' in shits):
    print('AE shit found')
    ParseAE()
if ('L2ifaces' in shits):
    print('L2ifaces shit found')
    ParseL2Ifaces()
if ('VRF' in shits):
    print('VRF shit found')
    ParseVRF()
if ('LIB' in shits):
    print('LIB shit found')
    ParseLIB()


for host in devices:
    fname = 'outputs/'+host+'.txt'
    thandler = open(fname, 'w')
    stuff = devices[host].PrintCfg()
    thandler.write(str(host))
    thandler.write(str(stuff))
    thandler.close()

